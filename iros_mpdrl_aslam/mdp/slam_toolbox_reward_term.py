from __future__ import annotations

import atexit
import csv
import math
import os
import time
from datetime import datetime

import numpy as np
import torch

from isaaclab.managers import ManagerTermBase, SceneEntityCfg
from isaaclab.managers.manager_term_cfg import RewardTermCfg
from isaaclab.sensors import Imu

from .observations import horizontal_scan
from .fixedlag_slam import FixedLagConfig, FixedLagSLAMVectorized


def _yaw_from_quat_wxyz(q: torch.Tensor) -> torch.Tensor:
    qw, qx, qy, qz = q.unbind(-1)
    siny_cosp = 2.0 * (qw * qz + qx * qy)
    cosy_cosp = 1.0 - 2.0 * (qy * qy + qz * qz)
    return torch.atan2(siny_cosp, cosy_cosp)


def _stable_logdet_3x3(Sigma: torch.Tensor, jitter: float = 1e-6) -> torch.Tensor:
    S = Sigma.to(torch.float32).clone()
    S[:, 0, 0] += float(jitter)
    S[:, 1, 1] += float(jitter)
    S[:, 2, 2] += float(jitter)

    a = S[:, 0, 0]
    b = S[:, 0, 1]
    c = S[:, 0, 2]
    d = S[:, 1, 0]
    e = S[:, 1, 1]
    f = S[:, 1, 2]
    g = S[:, 2, 0]
    h = S[:, 2, 1]
    i = S[:, 2, 2]

    det = a * (e * i - f * h) - b * (d * i - f * g) + c * (d * h - e * g)
    logdet = torch.log(torch.clamp(det, min=1e-30))
    logdet = torch.where(det > 0, logdet, torch.full_like(logdet, 50.0))
    return torch.nan_to_num(logdet, nan=50.0, posinf=50.0, neginf=-50.0)


def _scale_yaw_cov(Sigma: torch.Tensor, yaw_length_m: float) -> torch.Tensor:
    if yaw_length_m == 1.0:
        return Sigma
    S = Sigma.clone()
    L = float(yaw_length_m)
    S[:, 2, :] *= L
    S[:, :, 2] *= L
    return S


def _stable_logdet_3x3_np(Sigma: np.ndarray, jitter: float = 1e-6) -> float:
    S = np.asarray(Sigma, dtype=np.float64).reshape(3, 3).copy()
    S[0, 0] += float(jitter)
    S[1, 1] += float(jitter)
    S[2, 2] += float(jitter)

    a, b, c = S[0, 0], S[0, 1], S[0, 2]
    d, e, f = S[1, 0], S[1, 1], S[1, 2]
    g, h, i = S[2, 0], S[2, 1], S[2, 2]

    det = a * (e * i - f * h) - b * (d * i - f * g) + c * (d * h - e * g)
    if (not np.isfinite(det)) or det <= 0.0:
        return 50.0
    out = math.log(max(det, 1e-30))
    if not math.isfinite(out):
        return 50.0
    return float(max(min(out, 50.0), -50.0))


def _scale_yaw_cov_np(Sigma: np.ndarray, yaw_length_m: float) -> np.ndarray:
    if float(yaw_length_m) == 1.0:
        return np.asarray(Sigma, dtype=np.float64).reshape(3, 3)
    S = np.asarray(Sigma, dtype=np.float64).reshape(3, 3).copy()
    L = float(yaw_length_m)
    S[2, :] *= L
    S[:, 2] *= L
    return S


class SLAMTOOLBOXActiveReward(ManagerTermBase):
    """Active-SLAM reward.

    Modes:
      - default: original fixedlag uncertainty drives both reward and observation.
      - bridge mode: reward term synchronizes with slam_toolbox each step, computes the
        fixedlag->toolbox blend internally, and exposes the exact same scalar to the
        observation pipeline through env.unwrapped._slam_U_norm.

    This version can optionally require a *fresh* slam_toolbox covariance on every
    env step during the whole blend (`bridge_require_fresh_always=True`).
    """

    cfg: RewardTermCfg

    def __init__(self, cfg: RewardTermCfg, env):
        super().__init__(cfg, env)

        self._robot_name = cfg.params["robot_cfg"].name
        self._lidar_name = cfg.params["lidar_cfg"].name
        self._clear_map_on_reset = bool(cfg.params.get("clear_map_on_reset", True))

        slam_cfg_dict = dict(cfg.params.get("slam_cfg", {}))
        slam_cfg = FixedLagConfig(**slam_cfg_dict)

        device = env.device if isinstance(env.device, torch.device) else torch.device(env.device)
        self.slam = FixedLagSLAMVectorized(slam_cfg, num_envs=env.num_envs, device=device)

        N = env.num_envs
        H = int(cfg.params.get("visit_H", self.slam.cfg.map_H))
        W = int(cfg.params.get("visit_W", self.slam.cfg.map_W))
        self._visit_H = H
        self._visit_W = W
        self._visit_origin_x = float(cfg.params.get("visit_origin_x", self.slam.cfg.origin_x))
        self._visit_origin_y = float(cfg.params.get("visit_origin_y", self.slam.cfg.origin_y))
        self._visit_res = float(cfg.params.get("visit_resolution", self.slam.cfg.resolution))

        self._visit_epoch_map = torch.zeros((N, H, W), device=device, dtype=torch.int16)
        self._epoch_ctr = torch.zeros((N,), device=device, dtype=torch.int32)
        self._epoch_tag = torch.ones((N,), device=device, dtype=torch.int16)
        self._epoch_mod = max(1000, min(int(cfg.params.get("epoch_mod", 30000)), 32766))

        self._ep_step = torch.zeros((N,), device=device, dtype=torch.int32)
        self._visited_count = torch.zeros((N,), device=device, dtype=torch.float32)

        self._decay_steps = int(cfg.params.get("decay_steps", 2048))
        self._decay_alpha = float(max(0.0, min(float(cfg.params.get("decay_alpha", 0.5)), 1.0)))
        self._revisit_steps = int(cfg.params.get("revisit_steps", 0))

        self._visit_step_map: torch.Tensor | None = None
        if self._decay_steps > 0 or self._revisit_steps > 0:
            self._visit_step_map = torch.zeros((N, H, W), device=device, dtype=torch.int16)

        self._gate_visits_by_confidence = bool(cfg.params.get("gate_visits_by_confidence", False))

        self._prev_U = torch.zeros((N,), device=device, dtype=torch.float32)
        self._last_U = torch.zeros((N,), device=device, dtype=torch.float32)
        self._prev_U_raw = torch.zeros((N,), device=device, dtype=torch.float32)
        self._last_U_raw = torch.zeros((N,), device=device, dtype=torch.float32)
        self._last_fixedlag_U = torch.zeros((N,), device=device, dtype=torch.float32)

        self._U0_raw = torch.zeros((N,), device=device, dtype=torch.float32)
        self._U_ref_raw = torch.zeros((N,), device=device, dtype=torch.float32)
        self._U_hi_raw = torch.zeros((N,), device=device, dtype=torch.float32)

        self._u_span_min = max(float(cfg.params.get("u_raw_span", 6.0)), 1e-3)
        self._u_ref_update_eps = max(float(cfg.params.get("u_ref_update_eps", 0.02)), 0.0)
        self._u_hi_decay_beta = float(max(1e-5, min(float(cfg.params.get("u_hi_decay_beta", 0.01)), 0.2)))
        self._u_q_inflate = max(float(cfg.params.get("u_q_inflate", 0.0)), 0.0)

        self._wheel_radius = float(cfg.params.get("wheel_radius", 0.033))
        self._axle_length = float(cfg.params.get("axle_length", 0.160))
        self._left_joint_name = str(cfg.params.get("left_joint", "wheel_left_joint"))
        self._right_joint_name = str(cfg.params.get("right_joint", "wheel_right_joint"))

        self._wheel_left_id: int | None = None
        self._wheel_right_id: int | None = None
        self._wheel_odom_ready = False
        self._wheel_odom_warned = False
        self._try_init_wheel_odom()

        setattr(self._env.unwrapped, "_slam_dbg_step", 0)
        self._ar = torch.arange(N, device=device, dtype=torch.int64)

        self._init_metrics_logging()

        # ---------------- bridge state ----------------
        self._use_slam_toolbox_bridge = bool(cfg.params.get("use_slam_toolbox_bridge", False))
        self._bridge_mode = str(cfg.params.get("bridge_mode", "convex"))
        if self._bridge_mode not in {"convex", "inject_only"}:
            raise ValueError("bridge_mode must be 'convex' or 'inject_only'.")
        self._bridge_blend_steps = max(int(cfg.params.get("bridge_blend_steps", 204800)), 1)
        self._bridge_bootstrap_timeout_s = float(max(cfg.params.get("bridge_bootstrap_timeout_s", 20.0), 1e-3))
        self._bridge_step_wait_timeout_s = float(max(cfg.params.get("bridge_step_wait_timeout_s", 0.10), 1e-4))
        self._bridge_reset_quarantine_updates = max(int(cfg.params.get("bridge_reset_quarantine_updates", 2)), 0)
        self._bridge_require_fresh_when_full = bool(cfg.params.get("bridge_require_fresh_when_full", True))
        self._bridge_require_fresh_always = bool(cfg.params.get("bridge_require_fresh_always", False))
        if self._bridge_require_fresh_always and self._bridge_reset_quarantine_updates > 0:
            print(
                f"[INFO] bridge_require_fresh_always=True is incompatible with bridge_reset_quarantine_updates="
                f"{self._bridge_reset_quarantine_updates}. Forcing it to 0.",
                flush=True,
            )
            self._bridge_reset_quarantine_updates = 0
        self._bridge_total_env_steps = 0
        self._bridge_armed = torch.zeros((N,), device=device, dtype=torch.bool)
        self._bridge_last_odom_seq = torch.zeros((N,), device=device, dtype=torch.int64)
        self._bridge_seq_at_reset = torch.zeros((N,), device=device, dtype=torch.int64)
        self._bridge_allow_toolbox_after_seq = torch.zeros((N,), device=device, dtype=torch.int64)

        self._toolbox_ref_raw = np.full((N,), np.nan, dtype=np.float64)
        self._toolbox_hi_raw = np.full((N,), np.nan, dtype=np.float64)
        self._toolbox_last_raw = np.full((N,), np.nan, dtype=np.float64)
        self._toolbox_last_u = np.zeros((N,), dtype=np.float32)
        self._toolbox_has_valid = np.zeros((N,), dtype=bool)

    # ---------------- logging ----------------

    def _resolve_log_dir(self) -> str:
        p = self.cfg.params.get("slam_log_dir", None)
        if p:
            return os.path.abspath(str(p))
        for obj in (self._env, getattr(self._env, "unwrapped", None)):
            if obj is None:
                continue
            cfg = getattr(obj, "cfg", None)
            if cfg is not None and getattr(cfg, "log_dir", None):
                return os.path.abspath(str(cfg.log_dir))
            ld = getattr(obj, "log_dir", None)
            if ld:
                return os.path.abspath(str(ld))
        env_ld = os.environ.get("SLAM_LOG_DIR", "").strip()
        if env_ld:
            return os.path.abspath(env_ld)
        return os.path.abspath(os.getcwd())

    def _init_metrics_logging(self) -> None:
        def _env_flag(name: str, default: bool = False) -> bool:
            raw = os.environ.get(name, None)
            if raw is None:
                return default
            return str(raw).strip().lower() in ("1", "true", "yes", "y", "on")

        def _env_int(name: str, default: int) -> int:
            raw = os.environ.get(name, None)
            if raw is None or str(raw).strip() == "":
                return int(default)
            try:
                return int(raw)
            except Exception:
                return int(default)

        self._log_enabled = bool(self.cfg.params.get("log_slam_metrics", False)) or _env_flag("SLAM_LOG_METRICS", False)
        self._log_stride = max(1, int(self.cfg.params.get("slam_log_stride", _env_int("SLAM_LOG_STRIDE", 1))))
        self._log_flush_every = max(1, int(self.cfg.params.get("slam_log_flush_every", _env_int("SLAM_LOG_FLUSH_EVERY", 256))))
        self._log_write_xlsx = bool(self.cfg.params.get("slam_log_write_xlsx", False)) or _env_flag("SLAM_LOG_XLSX", False)

        self._log_dir = self._resolve_log_dir()
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        base = str(self.cfg.params.get("slam_log_filename", os.environ.get("SLAM_LOG_FILENAME", f"slam_metrics_{ts}")))
        if base.endswith(".csv"):
            base = base[:-4]
        if base.endswith(".xlsx"):
            base = base[:-5]
        self._log_csv_path = os.path.join(self._log_dir, base + ".csv")
        self._log_xlsx_path = os.path.join(self._log_dir, base + ".xlsx")

        self._log_fh = None
        self._log_writer = None
        self._log_buf: list[list[object]] = []
        self._log_global_step = 0
        self._log_finalized = False

        if not self._log_enabled:
            return

        try:
            os.makedirs(self._log_dir, exist_ok=True)
            self._open_log_file()
            try:
                setattr(self._env.unwrapped, "_slam_metrics_csv", self._log_csv_path)
                setattr(self._env.unwrapped, "_slam_metrics_xlsx", self._log_xlsx_path)
            except Exception:
                pass
            atexit.register(self._finalize_metrics_logging)
            print(f"[slam log] logging enabled -> {self._log_csv_path}", flush=True)
        except Exception as e:
            self._log_enabled = False
            print(f"[slam log] WARNING: failed to initialize metrics logging ({e})", flush=True)

    def _open_log_file(self) -> None:
        if self._log_fh is not None:
            return
        self._log_fh = open(self._log_csv_path, "w", newline="")
        self._log_writer = csv.writer(self._log_fh)
        self._log_writer.writerow(["global_step", "env_id", "episode_id", "episode_step", "U_norm", "rmse_xy_m", "mae_yaw_rad"])
        self._log_fh.flush()

    def _flush_log(self, force: bool = False) -> None:
        if (not self._log_enabled) or (self._log_fh is None) or (self._log_writer is None):
            return
        if len(self._log_buf) == 0:
            return
        if (not force) and (len(self._log_buf) < self._log_flush_every):
            return
        self._log_writer.writerows(self._log_buf)
        self._log_buf.clear()
        self._log_fh.flush()

    def _csv_to_xlsx(self, csv_path: str, xlsx_path: str) -> None:
        try:
            import openpyxl
            from openpyxl.styles import Font
            from openpyxl.utils import get_column_letter
        except Exception as e:
            print(f"[slam log] openpyxl unavailable, cannot write xlsx ({e}). CSV is still saved at: {csv_path}", flush=True)
            return

        def _coerce_value(v: str):
            if v is None:
                return None
            s = str(v).strip()
            if s == "":
                return ""
            try:
                if s.lstrip("-").isdigit():
                    return int(s)
                return float(s)
            except Exception:
                return v

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "slam_metrics"

        with open(csv_path, "r", newline="") as f:
            reader = csv.reader(f)
            for row_idx, row in enumerate(reader, start=1):
                if row_idx == 1:
                    ws.append(row)
                else:
                    ws.append([_coerce_value(x) for x in row])

        ws.freeze_panes = "A2"
        for cell in ws[1]:
            cell.font = Font(bold=True)
        widths = {1: 12, 2: 8, 3: 12, 4: 12, 5: 12, 6: 14, 7: 14}
        for col_idx, width in widths.items():
            ws.column_dimensions[get_column_letter(col_idx)].width = width
        wb.save(xlsx_path)

    def _finalize_metrics_logging(self) -> None:
        if (not self._log_enabled) or self._log_finalized:
            return
        self._log_finalized = True
        try:
            self._flush_log(force=True)
        except Exception as e:
            print(f"[slam log] WARNING: flush failed during finalize ({e})", flush=True)
        try:
            if self._log_fh is not None:
                self._log_fh.close()
        except Exception as e:
            print(f"[slam log] WARNING: close failed during finalize ({e})", flush=True)
        finally:
            self._log_fh = None
            self._log_writer = None
        if self._log_write_xlsx:
            try:
                self._csv_to_xlsx(self._log_csv_path, self._log_xlsx_path)
                print(f"[slam log] wrote: {self._log_xlsx_path}", flush=True)
            except Exception as e:
                print(f"[slam log] WARNING: xlsx conversion failed ({e})", flush=True)
        print(f"[slam log] wrote: {self._log_csv_path}", flush=True)

    def _log_metrics_step(self, U: torch.Tensor, rmse_xy: torch.Tensor, mae_yaw: torch.Tensor) -> None:
        if not self._log_enabled:
            return
        self._log_global_step += 1
        if (self._log_global_step % self._log_stride) != 0:
            return
        ep_id = self._epoch_ctr
        N = int(self._env.num_envs)
        if N == 1:
            self._log_buf.append([
                int(self._log_global_step),
                0,
                int(ep_id[0].item()),
                int(self._ep_step[0].item()),
                float(U[0].item()),
                float(rmse_xy[0].item()),
                float(mae_yaw[0].item()),
            ])
        else:
            for k in range(N):
                self._log_buf.append([
                    int(self._log_global_step),
                    int(k),
                    int(ep_id[k].item()),
                    int(self._ep_step[k].item()),
                    float(U[k].item()),
                    float(rmse_xy[k].item()),
                    float(mae_yaw[k].item()),
                ])
        self._flush_log(force=False)

    # ---------------- wheel odom ----------------

    def _try_init_wheel_odom(self) -> None:
        if self._wheel_odom_ready:
            return
        try:
            robot = self._env.scene[self._robot_name]
            l_ids, _ = robot.find_joints(self._left_joint_name)
            r_ids, _ = robot.find_joints(self._right_joint_name)
            self._wheel_left_id = int(l_ids[0])
            self._wheel_right_id = int(r_ids[0])
            self._wheel_odom_ready = True
        except Exception as e:
            self._wheel_odom_ready = False
            if not self._wheel_odom_warned:
                print(
                    f"[slam odom] WARNING: wheel encoder odom unavailable ({e}). Falling back to commanded v,w for SLAM.",
                    flush=True,
                )
                self._wheel_odom_warned = True

    def _wheel_odom_vw(self, robot) -> tuple[torch.Tensor | None, torch.Tensor | None]:
        if not self._wheel_odom_ready or self._wheel_left_id is None or self._wheel_right_id is None:
            return None, None
        jvel = getattr(robot.data, "joint_vel", None)
        if jvel is None or not isinstance(jvel, torch.Tensor) or jvel.ndim != 2:
            return None, None
        max_id = max(self._wheel_left_id, self._wheel_right_id)
        if jvel.shape[1] <= max_id:
            return None, None
        wL = jvel[:, self._wheel_left_id].to(torch.float32)
        wR = jvel[:, self._wheel_right_id].to(torch.float32)
        r = float(self._wheel_radius)
        L = float(max(self._axle_length, 1e-6))
        v = 0.5 * r * (wL + wR)
        w = (r / L) * (wR - wL)
        return v, w

    # ---------------- visitation ----------------

    def _xy_to_lin(self, xy_local: torch.Tensor) -> tuple[torch.Tensor, torch.Tensor]:
        x = xy_local[:, 0]
        y = xy_local[:, 1]
        gx = torch.floor((x - self._visit_origin_x) / self._visit_res).to(torch.int64)
        gy = torch.floor((y - self._visit_origin_y) / self._visit_res).to(torch.int64)
        valid = (gx >= 0) & (gx < self._visit_W) & (gy >= 0) & (gy < self._visit_H)
        gx_safe = torch.clamp(gx, 0, self._visit_W - 1)
        gy_safe = torch.clamp(gy, 0, self._visit_H - 1)
        lin = gy_safe * self._visit_W + gx_safe
        return lin, valid

    def _visit_reward_scale(self, lin: torch.Tensor, valid_for_update: torch.Tensor, valid_for_visit: torch.Tensor):
        N = self._env.num_envs
        ar = self._ar
        epoch_map_flat = self._visit_epoch_map.view(N, -1)
        prev_tag = epoch_map_flat[ar, lin]
        cur_tag = self._epoch_tag
        is_new_any = prev_tag != cur_tag

        new_update_mask = valid_for_update & is_new_any
        epoch_map_flat[ar[new_update_mask], lin[new_update_mask]] = cur_tag[ar[new_update_mask]]
        self._visited_count[new_update_mask] += 1.0

        new_mask = valid_for_visit & is_new_any
        scale = new_mask.to(torch.float32)
        revisit_mask = torch.zeros_like(new_mask)

        if self._visit_step_map is not None:
            step_map_flat = self._visit_step_map.view(N, -1)
            prev_step = step_map_flat[ar, lin].to(torch.int32)
            age = torch.clamp(self._ep_step - prev_step - 1, min=0)
            visited_this_ep = valid_for_visit & (~is_new_any)
            if self._decay_steps > 0:
                denom = float(max(self._decay_steps, 1))
                revisit_scale = (age.to(torch.float32) / denom).clamp(0.0, 1.0) * float(self._decay_alpha)
                revisit_scale = revisit_scale * visited_this_ep.to(torch.float32)
                revisit_mask = visited_this_ep & (age > 0) & (revisit_scale > 0.0)
                scale = scale + revisit_scale
            elif self._revisit_steps > 0:
                can_revisit = visited_this_ep & (age >= int(self._revisit_steps))
                revisit_mask = can_revisit
                scale = scale + can_revisit.to(torch.float32) * float(self._decay_alpha)
            idxv = ar[valid_for_update]
            linv = lin[valid_for_update]
            step_i16_all = torch.clamp(self._ep_step, 0, 32767).to(torch.int16)
            step_map_flat[idxv, linv] = step_i16_all[idxv]

        scale = torch.clamp(scale, 0.0, 1.0)
        return scale, new_mask, revisit_mask

    # ---------------- bridge helpers ----------------

    def _resolve_data_manager(self):
        return getattr(self._env.unwrapped, "data_manager", None)

    def _reset_toolbox_state(self, env_ids, baseline_seqs: list[int] | None = None):
        ids = [int(i) for i in env_ids]
        for i in ids:
            self._toolbox_ref_raw[i] = np.nan
            self._toolbox_hi_raw[i] = np.nan
            self._toolbox_last_raw[i] = np.nan
            self._toolbox_last_u[i] = 0.0
            self._toolbox_has_valid[i] = False
            base_seq = int(baseline_seqs[i]) if baseline_seqs is not None else 0
            self._bridge_seq_at_reset[i] = base_seq
            self._bridge_last_odom_seq[i] = base_seq
            self._bridge_allow_toolbox_after_seq[i] = base_seq + self._bridge_reset_quarantine_updates
            self._bridge_armed[i] = False

    def _bootstrap_toolbox_after_reset(self, env_ids):
        if not self._use_slam_toolbox_bridge:
            return
        dm = self._resolve_data_manager()
        if dm is None:
            return
        ids = [int(i) for i in env_ids]

        graph_timeout = min(float(self._bridge_bootstrap_timeout_s), 20.0)
        graph_ready, graph_status = dm.wait_for_toolbox_graph(ids, timeout_s=graph_timeout)
        if not graph_ready:
            raise RuntimeError(
                "slam_toolbox graph is not ready for the requested namespaces. "
                f"Graph status: {graph_status}. "
                "Check that you launched the same number of slam_toolbox instances as --num_envs, "
                "that the namespaces match (unitree_go2_0 ...), and that the nodes reached the active state."
            )

        cold_start = True
        snapshots_before = {i: dm.get_odom_seq_xy_yaw_cov(i) for i in ids}
        for i in ids:
            seq_i, _xy_i, _yaw_i, cov_i = snapshots_before[i]
            if int(seq_i) > 0 or cov_i is not None:
                cold_start = False
                break

        if not cold_start:
            dm.reset_slam(ids)
            time.sleep(0.10)

        baseline = dm._snapshot_odom_seqs()
        self._reset_toolbox_state(ids, baseline_seqs=baseline)
        return

    def _normalize_toolbox_uncertainty(self, env_idx: int, planar_cov, yaw_length_m: float):
        if planar_cov is None:
            return 0.0, float("nan"), False
        try:
            cov_scaled = _scale_yaw_cov_np(planar_cov, yaw_length_m=float(yaw_length_m))
            raw = _stable_logdet_3x3_np(cov_scaled)
        except Exception:
            return 0.0, float("nan"), False

        if not self._toolbox_has_valid[env_idx] or not np.isfinite(self._toolbox_ref_raw[env_idx]):
            ref = float(raw)
            hi = float(raw + self._u_span_min)
            u = 0.0
        else:
            ref = float(self._toolbox_ref_raw[env_idx])
            hi = float(self._toolbox_hi_raw[env_idx])
            if raw < (ref - self._u_ref_update_eps):
                ref = float(raw)
            hi_decay = (1.0 - self._u_hi_decay_beta) * hi + self._u_hi_decay_beta * float(raw)
            hi = float(max(float(raw), hi_decay))
            span = max(self._u_span_min, hi - ref)
            u = float(np.clip((float(raw) - ref) / span, 0.0, 1.0))

        self._toolbox_ref_raw[env_idx] = ref
        self._toolbox_hi_raw[env_idx] = hi
        self._toolbox_last_raw[env_idx] = float(raw)
        self._toolbox_last_u[env_idx] = float(u)
        self._toolbox_has_valid[env_idx] = True
        return float(u), float(raw), True

    def _step_toolbox(self, yaw_length_m: float, strict_full: bool = False):
        N = int(self._env.num_envs)
        device = self._env.device if isinstance(self._env.device, torch.device) else torch.device(self._env.device)
        dm = self._resolve_data_manager()
        if dm is None:
            return (
                torch.zeros((N,), device=device, dtype=torch.float32),
                torch.full((N,), float("nan"), device=device, dtype=torch.float32),
                torch.zeros((N,), device=device, dtype=torch.bool),
                torch.full((N, 3), float("nan"), device=device, dtype=torch.float32),
            )

        def _consume_snapshot(i: int, snapshot, scan_was_published: bool, last_before_local, fresh_mask, u, raw, pose):
            seq_i, xy_i, yaw_i, cov_i = snapshot
            seq_i = int(seq_i)

            if (not bool(self._bridge_armed[i].item())) and (seq_i > int(self._bridge_seq_at_reset[i].item())):
                self._bridge_armed[i] = True

            is_fresh = bool(scan_was_published) and (seq_i > int(last_before_local[i]))
            after_quarantine = seq_i >= int(self._bridge_allow_toolbox_after_seq[i].item())
            if is_fresh and after_quarantine and (cov_i is not None):
                ui, rawi, valid = self._normalize_toolbox_uncertainty(i, cov_i, yaw_length_m=yaw_length_m)
                if valid:
                    fresh_mask[i] = True
                    u[i] = ui
                    raw[i] = rawi
                    self._bridge_last_odom_seq[i] = seq_i
                    if xy_i is not None and yaw_i is not None:
                        pose[i, 0] = float(xy_i[0])
                        pose[i, 1] = float(xy_i[1])
                        pose[i, 2] = float(yaw_i)
            return seq_i, cov_i

        last_before = [int(x) for x in self._bridge_last_odom_seq.detach().cpu().tolist()]
        force_ids = [i for i in range(N) if not bool(self._bridge_armed[i].item())]
        scan_published = dm.pub_ros2_data(force_scan_envs=force_ids if force_ids else None)

        trigger_envs = [i for i in range(N) if bool(scan_published[i])]
        if trigger_envs:
            wait_budget = float(self._bridge_step_wait_timeout_s)
            if force_ids:
                wait_budget = max(wait_budget, min(float(self._bridge_bootstrap_timeout_s), 1.0))
            dm.wait_for_new_odom(trigger_envs, last_before, wait_budget)

        snapshots = [dm.get_odom_seq_xy_yaw_cov(i) for i in range(N)]

        fresh_mask = np.zeros((N,), dtype=bool)
        u = np.zeros((N,), dtype=np.float32)
        raw = np.full((N,), np.nan, dtype=np.float32)
        pose = np.full((N, 3), np.nan, dtype=np.float32)

        for i in range(N):
            _consume_snapshot(i, snapshots[i], bool(scan_published[i]), last_before, fresh_mask, u, raw, pose)

        rescue_ids = [i for i in force_ids if not fresh_mask[i]]
        if strict_full:
            rescue_ids = sorted(set(rescue_ids + [i for i in range(N) if not fresh_mask[i]]))

        if rescue_ids:
            deadline = time.time() + max(float(self._bridge_step_wait_timeout_s), min(float(self._bridge_bootstrap_timeout_s), 2.0))
            while rescue_ids and time.time() < deadline:
                dm.pub_ros2_data(force_scan_envs=rescue_ids)
                remaining = max(0.0, deadline - time.time())
                dm.wait_for_new_odom(rescue_ids, last_before, min(0.05, remaining))
                new_rescue_ids = []
                for i in rescue_ids:
                    snapshot_i = dm.get_odom_seq_xy_yaw_cov(i)
                    seq_i, _cov_i = _consume_snapshot(i, snapshot_i, True, last_before, fresh_mask, u, raw, pose)
                    if not fresh_mask[i]:
                        new_rescue_ids.append(i)
                rescue_ids = new_rescue_ids

        return (
            torch.as_tensor(u, device=device, dtype=torch.float32),
            torch.as_tensor(raw, device=device, dtype=torch.float32),
            torch.as_tensor(fresh_mask, device=device, dtype=torch.bool),
            torch.as_tensor(pose, device=device, dtype=torch.float32),
        )

    # ---------------- pose helpers ----------------

    def _compute_gt_pose(self, env, robot):
        root = robot.data.root_state_w
        xy_local = root[:, 0:2] - env.scene.env_origins[:, :2]
        yaw_gt = _yaw_from_quat_wxyz(root[:, 3:7])
        gt = torch.stack([xy_local[:, 0], xy_local[:, 1], yaw_gt], dim=-1)
        return gt, xy_local

    def _compute_pose_errors(self, est: torch.Tensor, gt: torch.Tensor):
        err = gt - est
        err[:, 2] = torch.atan2(torch.sin(err[:, 2]), torch.cos(err[:, 2]))
        rmse_xy = torch.sqrt(torch.sum(err[:, :2] ** 2, dim=-1))
        mae_yaw = torch.abs(err[:, 2])
        return rmse_xy, mae_yaw

    def _compute_fixedlag_signal(self, env, ranges, imu_wz, v_odom, w_odom, yaw_length_m: float, gt_pose: torch.Tensor):
        _mu, _Sigma = self.slam.step(v_odom, w_odom, float(env.step_dt), ranges, imu_wz=imu_wz)

        Sigma_pre = getattr(self.slam, "Sigma_pre", None)
        if Sigma_pre is None:
            Sigma_pre = getattr(self.slam, "Sigma", None)

        Sigma_f = _scale_yaw_cov(Sigma_pre.to(torch.float32), yaw_length_m=float(yaw_length_m))
        U_raw = _stable_logdet_3x3(Sigma_f, jitter=1e-6)
        U_raw = torch.nan_to_num(U_raw, nan=50.0, posinf=50.0, neginf=-50.0)

        q = getattr(self.slam, "dbg_meas_q", None)
        if q is None:
            q = torch.zeros_like(U_raw)
        else:
            q = torch.nan_to_num(q.to(U_raw.dtype), nan=0.0, posinf=0.0, neginf=0.0).clamp(0.0, 1.0)

        U_raw_eff = U_raw + float(self._u_q_inflate) * (1.0 - q)

        ref_cur = self._U_ref_raw
        ref_new = torch.where(U_raw_eff < (ref_cur - self._u_ref_update_eps), U_raw_eff, ref_cur)
        self._U_ref_raw = ref_new

        hi_cur = self._U_hi_raw
        hi_decay = (1.0 - self._u_hi_decay_beta) * hi_cur + self._u_hi_decay_beta * U_raw_eff
        hi_new = torch.maximum(U_raw_eff, hi_decay)
        self._U_hi_raw = hi_new

        span = torch.clamp(self._U_hi_raw - self._U_ref_raw, min=self._u_span_min)
        U = (U_raw_eff - self._U_ref_raw) / span
        U = torch.nan_to_num(U, nan=0.0, posinf=1.0, neginf=0.0).clamp(0.0, 1.0)

        est = self.slam.mu_pre
        rmse_xy, mae_yaw = self._compute_pose_errors(est, gt_pose)
        return U, U_raw_eff, q, rmse_xy, mae_yaw

    # ---------------- reset ----------------

    def reset(self, env_ids=None):
        if env_ids is None or isinstance(env_ids, slice):
            env_ids = torch.arange(self._env.num_envs, device=self._env.device)
        else:
            env_ids = torch.as_tensor(env_ids, device=self._env.device, dtype=torch.long)

        robot = self._env.scene[self._robot_name]
        root = robot.data.root_state_w[env_ids]
        xy = root[:, 0:2]
        yaw = _yaw_from_quat_wxyz(root[:, 3:7])
        origins = self._env.scene.env_origins[env_ids, :2]
        xy_local = xy - origins
        init_pose = torch.stack([xy_local[:, 0], xy_local[:, 1], yaw], dim=-1)

        self.slam.reset(env_ids=env_ids, init_pose_xyyaw=init_pose, clear_map=self._clear_map_on_reset)

        self._epoch_ctr[env_ids] += 1
        new_tag = (self._epoch_ctr[env_ids] % self._epoch_mod) + 1
        new_tag_i16 = new_tag.to(torch.int16)
        wrapped = new_tag_i16 == 1
        if wrapped.any():
            over_ids = env_ids[wrapped]
            self._visit_epoch_map[over_ids] = 0
            if self._visit_step_map is not None:
                self._visit_step_map[over_ids] = 0

        self._epoch_tag[env_ids] = new_tag_i16
        self._ep_step[env_ids] = 0
        self._visited_count[env_ids] = 0.0

        lin0, valid0 = self._xy_to_lin(xy_local)
        if valid0.any():
            epoch_map_flat = self._visit_epoch_map.view(self._env.num_envs, -1)
            eid = env_ids[valid0]
            linm = lin0[valid0]
            epoch_map_flat[eid, linm] = self._epoch_tag[eid]
            if self._visit_step_map is not None:
                step_map_flat = self._visit_step_map.view(self._env.num_envs, -1)
                step_map_flat[eid, linm] = torch.zeros_like(self._epoch_tag[eid], dtype=torch.int16)
            self._visited_count[eid] = 1.0

        yaw_length_m = float(self.cfg.params.get("yaw_length_m", 1.0))
        Sigma_pre = getattr(self.slam, "Sigma_pre", None)
        if Sigma_pre is None:
            Sigma_pre = getattr(self.slam, "Sigma", None)
        Sigma0 = _scale_yaw_cov(Sigma_pre[env_ids].to(torch.float32), yaw_length_m=yaw_length_m)
        U0_raw = _stable_logdet_3x3(Sigma0, jitter=1e-6)
        U0_raw = torch.nan_to_num(U0_raw, nan=50.0, posinf=50.0, neginf=-50.0)

        self._U0_raw[env_ids] = U0_raw
        self._U_ref_raw[env_ids] = U0_raw
        self._U_hi_raw[env_ids] = U0_raw + float(self._u_span_min)

        self._prev_U[env_ids] = 0.0
        self._last_U[env_ids] = 0.0
        self._prev_U_raw[env_ids] = U0_raw
        self._last_U_raw[env_ids] = U0_raw
        self._last_fixedlag_U[env_ids] = 0.0

        setattr(self._env.unwrapped, "_slam_U_norm", self._last_U)
        setattr(self._env.unwrapped, "_slam_U", self._last_U)
        setattr(self._env.unwrapped, "_slam_U_raw", self._last_U_raw)
        setattr(self._env.unwrapped, "_slam_U_fixedlag_norm", self._last_fixedlag_U)
        setattr(self._env.unwrapped, "_slam_toolbox_U_norm", torch.zeros_like(self._last_U))
        setattr(self._env.unwrapped, "_slam_toolbox_U_raw", torch.full_like(self._last_U, float("nan")))
        setattr(self._env.unwrapped, "_slam_bridge_alpha", float(min(1.0, self._bridge_total_env_steps / float(self._bridge_blend_steps))))
        setattr(self._env.unwrapped, "_slam_dbg_step", 0)

        if self._use_slam_toolbox_bridge:
            self._bootstrap_toolbox_after_reset(env_ids.detach().cpu().tolist())

    # ---------------- main reward call ----------------

    @torch.no_grad()
    def __call__(
        self,
        env,
        robot_cfg: SceneEntityCfg,
        lidar_cfg: SceneEntityCfg,
        imu_cfg: SceneEntityCfg | None = None,
        offset: float = 0.0,
        max_range: float = 10.0,
        yaw_length_m: float = 1.0,
        u_hi: float = 0.7,
        explore_w: float = 3.0,
        dU_scale: float = 0.05,
        conf_power: float = 2.0,
        loop_abs_w: float = 0.2,
        loop_delta_w: float = 1.5,
        lidar_noise_std: float = 0.00,
        lidar_clip_min: float = 0.1,
        reward_clip: float = 100.0,
        print_every: int = 0,
        k_exp: float = 2.0,
        u_q_inflate: float | None = None,
        u_hi_decay_beta: float | None = None,
        u_ref_update_eps: float | None = None,
        u_span_min: float | None = None,
        use_slam_toolbox_bridge: bool | None = None,
        bridge_blend_steps: int | None = None,
        bridge_mode: str | None = None,
        bridge_bootstrap_timeout_s: float | None = None,
        bridge_step_wait_timeout_s: float | None = None,
        bridge_reset_quarantine_updates: int | None = None,
        bridge_require_fresh_when_full: bool | None = None,
        bridge_require_fresh_always: bool | None = None,
    ) -> torch.Tensor:
        if not hasattr(self, "_once"):
            self._once = True
            print("SLAMTOOLBOXActiveReward __call__ is running", flush=True)

        if u_q_inflate is not None:
            self._u_q_inflate = max(float(u_q_inflate), 0.0)
        if u_hi_decay_beta is not None:
            self._u_hi_decay_beta = float(max(1e-5, min(float(u_hi_decay_beta), 0.2)))
        if u_ref_update_eps is not None:
            self._u_ref_update_eps = max(float(u_ref_update_eps), 0.0)
        if u_span_min is not None:
            self._u_span_min = max(float(u_span_min), 1e-3)

        if use_slam_toolbox_bridge is not None:
            self._use_slam_toolbox_bridge = bool(use_slam_toolbox_bridge)
        if bridge_blend_steps is not None:
            self._bridge_blend_steps = max(int(bridge_blend_steps), 1)
        if bridge_mode is not None:
            bridge_mode = str(bridge_mode)
            if bridge_mode not in {"convex", "inject_only"}:
                raise ValueError("bridge_mode must be 'convex' or 'inject_only'.")
            self._bridge_mode = bridge_mode
        if bridge_bootstrap_timeout_s is not None:
            self._bridge_bootstrap_timeout_s = float(max(float(bridge_bootstrap_timeout_s), 1e-3))
        if bridge_step_wait_timeout_s is not None:
            self._bridge_step_wait_timeout_s = float(max(float(bridge_step_wait_timeout_s), 1e-4))
        if bridge_reset_quarantine_updates is not None:
            self._bridge_reset_quarantine_updates = max(int(bridge_reset_quarantine_updates), 0)
        if bridge_require_fresh_when_full is not None:
            self._bridge_require_fresh_when_full = bool(bridge_require_fresh_when_full)
        if bridge_require_fresh_always is not None:
            self._bridge_require_fresh_always = bool(bridge_require_fresh_always)
        if self._bridge_require_fresh_always and self._bridge_reset_quarantine_updates > 0:
            self._bridge_reset_quarantine_updates = 0

        self._ep_step += 1

        vw = getattr(env.unwrapped, "_last_cmd_vw", None)
        if vw is not None:
            v_cmd = vw[:, 0]
            w_cmd = vw[:, 1]
        else:
            w = getattr(env.unwrapped, "_last_wheel_cmd", None)
            if w is None:
                a = env.action_manager.action
                v_cmd = a[:, 0]
                w_cmd = a[:, 1]
            else:
                r, L = 0.033, 0.160
                wL, wR = w[:, 0], w[:, 1]
                v_cmd = 0.5 * r * (wL + wR)
                w_cmd = (r / L) * (wR - wL)

        robot = env.scene[self._robot_name]
        if not self._wheel_odom_ready:
            self._try_init_wheel_odom()
        v_odom, w_odom = self._wheel_odom_vw(robot)
        if v_odom is None or w_odom is None:
            v_odom = v_cmd
            w_odom = w_cmd

        ranges = horizontal_scan(env, lidar_cfg, offset=offset, max_range=max_range)
        if lidar_noise_std > 0.0:
            ranges = ranges + float(lidar_noise_std) * torch.randn_like(ranges)
        ranges = torch.clamp(ranges, float(lidar_clip_min), float(max_range))

        imu_wz = None
        if imu_cfg is not None and imu_cfg.name in env.scene.sensors:
            imu: Imu = env.scene.sensors[imu_cfg.name]
            imu_wz = imu.data.ang_vel_b[:, 2]

        gt_pose, xy_local = self._compute_gt_pose(env, robot)

        alpha = 0.0
        if self._use_slam_toolbox_bridge:
            alpha = float(min(1.0, self._bridge_total_env_steps / float(self._bridge_blend_steps)))

        toolbox_u = torch.zeros((env.num_envs,), device=env.device, dtype=torch.float32)
        toolbox_raw = torch.full((env.num_envs,), float("nan"), device=env.device, dtype=torch.float32)
        toolbox_fresh = torch.zeros((env.num_envs,), device=env.device, dtype=torch.bool)
        toolbox_pose = torch.full((env.num_envs, 3), float("nan"), device=env.device, dtype=torch.float32)
        strict_bridge = False
        if self._use_slam_toolbox_bridge:
            strict_bridge = bool(
                self._bridge_require_fresh_always
                or (alpha >= 1.0 - 1e-8 and self._bridge_require_fresh_when_full)
            )
            toolbox_u, toolbox_raw, toolbox_fresh, toolbox_pose = self._step_toolbox(
                yaw_length_m=float(yaw_length_m),
                strict_full=strict_bridge,
            )

        if strict_bridge:
            missing = (~toolbox_fresh).nonzero(as_tuple=False).flatten()
            if missing.numel() > 0:
                raise RuntimeError(
                    "slam_toolbox did not deliver a fresh covariance for every env step. "
                    f"Missing env ids: {missing.detach().cpu().tolist()}. "
                    "Keep fresh-per-step training by lowering slam_toolbox motion/time thresholds or increasing the per-step wait timeout."
                )

        need_fixedlag = True
        if self._use_slam_toolbox_bridge and bool(toolbox_fresh.all().item()) and alpha >= 1.0 - 1e-8:
            need_fixedlag = False

        if need_fixedlag:
            U_fixed, U_fixed_raw_eff, q, rmse_xy, mae_yaw = self._compute_fixedlag_signal(
                env, ranges, imu_wz, v_odom, w_odom, yaw_length_m=float(yaw_length_m), gt_pose=gt_pose
            )
            self._last_fixedlag_U = U_fixed.detach()
        else:
            U_fixed = self._last_fixedlag_U.detach().clone()
            U_fixed_raw_eff = torch.full((env.num_envs,), float("nan"), device=env.device, dtype=torch.float32)
            q = torch.ones((env.num_envs,), device=env.device, dtype=torch.float32)
            rmse_xy, mae_yaw = self._compute_pose_errors(toolbox_pose, gt_pose)

        if self._use_slam_toolbox_bridge:
            if self._bridge_mode == "inject_only":
                U = U_fixed.clone()
                injected = torch.clamp(alpha * toolbox_u, 0.0, 1.0)
                U = torch.where(toolbox_fresh, injected, U_fixed)
            else:
                blend = torch.clamp((1.0 - alpha) * U_fixed + alpha * toolbox_u, 0.0, 1.0)
                U = torch.where(toolbox_fresh, blend, U_fixed)
        else:
            U = U_fixed

        env.unwrapped._slam_U_fixedlag_norm = self._last_fixedlag_U.detach().view(-1)
        env.unwrapped._slam_toolbox_U_norm = toolbox_u.detach().view(-1)
        env.unwrapped._slam_toolbox_U_raw = toolbox_raw.detach().view(-1)
        env.unwrapped._slam_bridge_alpha = float(alpha)
        env.unwrapped._slam_U_norm = U.detach().view(-1)

        U_prev = self._prev_U
        dU_norm = U_prev - U

        self._log_metrics_step(U, rmse_xy, mae_yaw)

        lin, valid = self._xy_to_lin(xy_local)
        valid_for_update = valid
        valid_for_visit = valid
        if self._gate_visits_by_confidence:
            conf_gate = (U_prev < float(u_hi))
            valid_for_update = valid_for_update & conf_gate
            valid_for_visit = valid_for_visit & conf_gate

        visit_scale, new_cell_mask, revisit_mask = self._visit_reward_scale(
            lin, valid_for_update=valid_for_update, valid_for_visit=valid_for_visit
        )
        r_explore = float(explore_w) * visit_scale

        u_hi_val = float(u_hi)
        conf = torch.clamp((u_hi_val - U_prev) / max(u_hi_val, 1e-6), 0.0, 1.0)
        conf = conf.pow(float(max(1.0, conf_power)))

        dU_pos = torch.clamp(dU_norm, min=0.0)
        info = torch.clamp(dU_pos / max(float(dU_scale), 1e-6), 0.0, 1.0)

        mult = conf + (1.0 - conf) * info
        r_explore_term = mult * r_explore

        min_beam = ranges.amin(dim=1)
        d_crash = 0.3
        d_full = 0.5
        hug_walls_pen = torch.clamp((min_beam - d_crash) / (d_full - d_crash), 0.0, 1.0)
        r_explore_term *= hug_walls_pen

        w = torch.exp(float(k_exp) * U_prev)
        r_unc_state = -float(loop_abs_w) * w * U_prev
        r_unc_delta = float(loop_delta_w) * w * (U_prev - U)
        r_unc_delta = torch.where(r_unc_delta > 0.0, r_unc_delta * hug_walls_pen, r_unc_delta)

        r = r_explore_term + r_unc_state + r_unc_delta
        r = torch.nan_to_num(r, nan=-float(reward_clip), posinf=float(reward_clip), neginf=-float(reward_clip))
        r = torch.clamp(r, -float(reward_clip), float(reward_clip))

        self._prev_U = U.detach()
        self._last_U = U.detach()
        self._prev_U_raw = U_fixed_raw_eff.detach()
        self._last_U_raw = U_fixed_raw_eff.detach()

        setattr(env.unwrapped, "_slam_U", self._last_U)
        setattr(env.unwrapped, "_slam_U_raw", self._last_U_raw)
        setattr(env.unwrapped, "_slam_bridge_external_ready", toolbox_fresh.detach().view(-1))

        step = int(getattr(env.unwrapped, "_slam_dbg_step", 0)) + 1
        setattr(env.unwrapped, "_slam_dbg_step", step)

        if print_every is not None and int(print_every) > 0 and (step % int(print_every) == 0):
            def qsafe(x: torch.Tensor, p: float) -> float:
                x = x.detach().float().flatten()
                x = x[torch.isfinite(x)]
                if x.numel() == 0:
                    return float("nan")
                return torch.quantile(x, p).item()

            Rc = r.detach().float()
            print(
                "[SLAM/U dbg]\n"
                f"  U_used: mean={U.mean().item():.3f} p50={qsafe(U,0.50):.3f} p90={qsafe(U,0.90):.3f}\n"
                f"  U_fixedlag: mean={self._last_fixedlag_U.mean().item():.3f}  toolbox(mean)={toolbox_u.mean().item():.3f}  alpha={alpha:.3f}\n"
                f"  toolbox fresh frac={toolbox_fresh.float().mean().item():.3f}  q(fixedlag) mean={q.mean().item():.3f}\n"
                f"  pose err: rmse_xy mean={rmse_xy.mean().item():.3f}  mae_yaw mean={mae_yaw.mean().item():.3f}\n"
                f"  visit: new_frac={new_cell_mask.float().mean().item():.3f} revisit_frac={revisit_mask.float().mean().item():.3f} unique_cells={self._visited_count.mean().item():.1f}\n"
                f"  reward parts(mean): explore={r_explore_term.mean().item():.3f} unc_state={r_unc_state.mean().item():.3f} unc_delta={r_unc_delta.mean().item():.3f} total={Rc.mean().item():.3f}",
                flush=True,
            )

        if self._use_slam_toolbox_bridge:
            self._bridge_total_env_steps += int(toolbox_fresh.to(torch.int32).sum().item())

        return r / float(env.step_dt)